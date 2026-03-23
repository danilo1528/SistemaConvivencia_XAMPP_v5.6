"""
excel_generator.py — Generador Excel
Instrumento No. 001 — Tarjeta de Deméritos
Instrumento No. 002 — Registro Consolidado Mensual
Memorándum N.° 06-2025 MINEDUCYT
"""
import io, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

MESES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
             7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

CAUSALES = {
    "A":"A) No saludar al entrar o al salir del aula.",
    "B":"B) Omitir «Por favor» al hacer una petición.",
    "C":"C) Omitir «Gracias» al recibir un favor, material o atención.",
    "D":"D) Usar un tono grosero o irrespetuoso.",
}
OPCIONES_RED = {
    "A":"A) Semana completa con saludos y cortesía ejemplares.",
    "B":"B) Apoyar en actividades de orden y limpieza escolar.",
    "C":"C) Participar en campañas de valores.",
}
TIPOS_REC = {
    "A":"A) Diploma de Mención Honorífica de Cortesía Escolar.",
    "B":"B) Mención en Mural Escolar.",
}
ESCALA = [
    (3,5,"Advertencia verbal y reflexión escrita"),
    (6,9,"Comunicación a familia y tarea correctiva"),
    (10,10,"Suspensión de privilegios escolares"),
    (11,14,"Reunión con dirección y familia"),
    (15,99,"El estudiante NO podrá ser promovido de grado"),
]

# Paleta
AZ="003876"; AZL="1565C0"; AZP="DDEEFF"; GRP="ECEFF1"
BLA="FFFFFF"; NEG="000000"; ROJ="C62828"; ROJP="FFEBEE"
VER="1B5E20"; VERP="E8F5E9"; ORA="E65100"; ORAP="FFF3E0"

def _s(color="CCCCCC",style="thin"):
    return Side(style=style,color=color)

def _border(color="BBBBBB",style="thin"):
    s=_s(color,style); return Border(left=s,right=s,top=s,bottom=s)

def _fill(c): return PatternFill("solid",fgColor=c)
def _font(bold=False,color=NEG,size=10,name="Arial"):
    return Font(bold=bold,color=color,size=size,name=name)
def _align(h="left",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

def _hdr(ws,row,sc,ec,txt,bg=AZ,fg=BLA,size=10,bold=True):
    """Merge sc:ec and write header. Safe — only writes to top-left cell."""
    if sc < ec:
        ws.merge_cells(start_row=row,start_column=sc,end_row=row,end_column=ec)
    c = ws.cell(row=row,column=sc)
    if isinstance(c,MergedCell): return
    c.value=txt
    c.font=_font(bold,fg,size)
    c.fill=_fill(bg)
    c.alignment=_align("center","center",True)
    c.border=_border(AZ,"medium")

def _cell(ws,row,col,value="",bg=BLA,fg=NEG,bold=False,size=9,
          halign="left",wrap=False,bc="BBBBBB"):
    c=ws.cell(row=row,column=col)
    if isinstance(c,MergedCell): return c
    c.value=value
    c.font=_font(bold,fg,size)
    c.fill=_fill(bg)
    c.alignment=_align(halign,"center",wrap)
    c.border=_border(bc)
    return c


class ExcelGenerator:
    def __init__(self,output_dir="reports"):
        os.makedirs(output_dir,exist_ok=True)
        self.output_dir=output_dir

    # ── INSTRUMENTO 001 ──────────────────────────────────────────────────────
    def generar_tarjeta_demerito(self,estudiante,registros,totales,mes,anio,cfg):
        wb=Workbook(); ws=wb.active; ws.title="Instrumento No. 001"
        cw=[4,11,7,7,7,7,7,7,7,14,14,10,8]
        for i,w in enumerate(cw,1): ws.column_dimensions[get_column_letter(i)].width=w
        L=13; row=1

        # Títulos
        _hdr(ws,row,1,L,"MINISTERIO DE EDUCACIÓN, CIENCIA Y TECNOLOGÍA — MINEDUCYT",AZ,BLA,12); row+=1
        _hdr(ws,row,1,L,"Instrumento No. 001 — Tarjeta de Deméritos del Estudiante",AZL,BLA,11); row+=1
        _hdr(ws,row,1,L,"Memorándum N.° 06-2025 — Reglamento para la Promoción de la Cortesía Escolar",AZP,AZ,9,False); row+=1

        # CE
        _hdr(ws,row,1,L,"DATOS DEL CENTRO EDUCATIVO (Campos 1-5)",AZ,BLA,10); row+=1
        ce=[("1. Centro Educativo:",cfg.get("NOMBRE_INSTITUCION","—")),
            ("2. Código C.E.:",cfg.get("CODIGO_CE","—")),
            ("3. Departamento:",cfg.get("DEPARTAMENTO_CE","—")),
            ("4. Municipio:",cfg.get("MUNICIPIO_CE","—")),
            ("5. Distrito:",cfg.get("DISTRITO_CE","—"))]
        col=1
        for i,(lbl,val) in enumerate(ce):
            if i>0 and i%3==0: row+=1; col=1
            _cell(ws,row,col,lbl,AZP,AZ,True,8,"right"); col+=1
            span=6 if i==0 else 2
            ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
            _cell(ws,row,col,val,BLA,NEG,False,9); col+=span+1
        row+=1

        # Estudiante
        _hdr(ws,row,1,L,"DATOS DEL ESTUDIANTE (Campos 6-10)",AZ,BLA,10); row+=1
        nombre_c=f"{estudiante.get('nombre','')} {estudiante.get('apellido','')}".strip()
        grado_s=f"{estudiante.get('grado_nombre','')} {estudiante.get('seccion','')}".strip()
        row1=[("6. Nombre:",nombre_c,5),("7. NIE:",estudiante.get("nie","—"),1),
              ("8. Sexo:","Mujer (M)" if estudiante.get("sexo")=="M" else "Hombre (H)",1)]
        col=1
        for lbl,val,span in row1:
            _cell(ws,row,col,lbl,AZP,AZ,True,9,"right"); col+=1
            ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
            _cell(ws,row,col,val,BLA,NEG,False,9); col+=span
        row+=1
        row2=[("9. Grado/Sección:",grado_s,3),("10. Turno:",estudiante.get("turno","—"),2),
              ("Mes/Año:",f"{MESES_ES.get(mes,mes)}/{anio}",2)]
        col=1
        for lbl,val,span in row2:
            _cell(ws,row,col,lbl,AZP,AZ,True,9,"right"); col+=1
            ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
            _cell(ws,row,col,val,BLA,NEG,False,9); col+=span
        row+=1

        # Tabla referencia
        _hdr(ws,row,1,L,"TABLA DE REFERENCIA — REGLAMENTO DE CORTESÍA ESCOLAR",GRP,AZ,9); row+=1
        for sc,ec,txt,bg in [(1,4,"Art. 3 — Causales",AZL),(5,8,"Art. 6 — Redenciones",AZL),
                              (9,11,"Art. 7 — Reconocimientos",AZL),(12,13,"Art. 5 — Consecuencias",AZL)]:
            _hdr(ws,row,sc,ec,txt,bg,BLA,8)
        row+=1
        ci=list(CAUSALES.items()); oi=list(OPCIONES_RED.items())
        ri=list(TIPOS_REC.items()); ei=[(str(a)+("-"+str(b) if b<99 else "+"),c) for a,b,c in ESCALA]
        for i in range(max(len(ci),len(oi),len(ri),len(ei))):
            bg2=AZP if i%2==0 else BLA
            v=ci[i][1][3:] if i<len(ci) else ""
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
            _cell(ws,row,1,v,bg2,NEG,False,8,wrap=True)
            v=oi[i][1][3:] if i<len(oi) else ""
            ws.merge_cells(start_row=row,start_column=5,end_row=row,end_column=8)
            _cell(ws,row,5,v,bg2,NEG,False,8,wrap=True)
            v=ri[i][1][3:] if i<len(ri) else ""
            ws.merge_cells(start_row=row,start_column=9,end_row=row,end_column=11)
            _cell(ws,row,9,v,VERP if i%2==0 else BLA,NEG,False,8,wrap=True)
            if i<len(ei):
                _cell(ws,row,12,ei[i][0],ORAP if i%2==0 else BLA,NEG,True,8,"center")
                _cell(ws,row,13,ei[i][1],ORAP if i%2==0 else BLA,NEG,False,8,wrap=True)
            ws.row_dimensions[row].height=28; row+=1

        # Encabezados tabla registros
        _hdr(ws,row,1,L,"REGISTRO DE DEMÉRITOS, REDENCIONES Y RECONOCIMIENTOS",AZ,BLA,10); row+=1
        for sc,ec,txt in [(1,1,"No."),(2,2,"11. Fecha"),(3,6,"12. D — Deméritos"),
                           (7,9,"13. R — Redención"),(10,11,"14. RC — Reconocim."),
                           (12,12,"15. Docente"),(13,13,"17. Firma")]:
            _hdr(ws,row,sc,ec,txt,AZL,BLA,8)
        row+=1
        _cell(ws,row,1,"",AZP); _cell(ws,row,2,"Fecha",AZP,AZ,True,8,"center")
        for ci2,l in enumerate(["A","B","C","D"],3): _cell(ws,row,ci2,f"D:{l}",ROJP,ROJ,True,9,"center")
        for ci2,l in enumerate(["A","B","C"],7):     _cell(ws,row,ci2,f"R:{l}",VERP,VER,True,9,"center")
        for ci2,l in enumerate(["A","B"],10):        _cell(ws,row,ci2,f"RC:{l}",ORAP,ORA,True,9,"center")
        _cell(ws,row,12,"Nombre",AZP,AZ,True,8,"center")
        _cell(ws,row,13,"Firma",AZP,AZ,True,8,"center")
        row+=1

        # Filas de datos
        nf=max(8,len(registros))
        for i in range(nf):
            reg=registros[i] if i<len(registros) else {}
            bg3=GRP if i%2==0 else BLA
            _cell(ws,row,1,i+1,AZP,AZ,True,9,"center")
            fecha=str(reg.get("fecha","") or "")
            if len(fecha)==10 and "-" in fecha:
                p=fecha.split("-"); fecha=f"{p[2]}/{p[1]}/{p[0]}"
            _cell(ws,row,2,fecha,bg3,NEG,False,9,"center")
            for ci2,cod in enumerate(["A","B","C","D"],3):
                v="✓" if reg.get("causal_demerito")==cod else ""
                _cell(ws,row,ci2,v,ROJP if v else bg3,ROJ if v else NEG,bool(v),11,"center")
            for ci2,cod in enumerate(["A","B","C"],7):
                v="✓" if reg.get("opcion_redencion")==cod else ""
                _cell(ws,row,ci2,v,VERP if v else bg3,VER if v else NEG,bool(v),11,"center")
            for ci2,cod in enumerate(["A","B"],10):
                v="✓" if reg.get("tipo_reconocimiento")==cod else ""
                _cell(ws,row,ci2,v,ORAP if v else bg3,ORA if v else NEG,bool(v),11,"center")
            _cell(ws,row,12,str(reg.get("docente_nombre","") or "")[:20],bg3,NEG,False,8,wrap=True)
            _cell(ws,row,13,"✓" if reg.get("firma_estudiante") else "",bg3,VER,False,11,"center")
            row+=1

        # Totales fila 18
        _cell(ws,row,1,"18.",AZ,BLA,True,9,"center")
        _cell(ws,row,2,"TOTAL",AZ,BLA,True,9,"center")
        for ci2,k in enumerate(["d_A","d_B","d_C","d_D"],3):
            _cell(ws,row,ci2,int(totales.get(k,0) or 0),ROJ,BLA,True,11,"center")
        for ci2,k in enumerate(["r_A","r_B","r_C"],7):
            _cell(ws,row,ci2,int(totales.get(k,0) or 0),VER,BLA,True,11,"center")
        for ci2,k in enumerate(["rc_A","rc_B"],10):
            _cell(ws,row,ci2,int(totales.get(k,0) or 0),ORA,BLA,True,11,"center")
        saldo=int(totales.get("saldo_neto",0) or 0)
        ws.merge_cells(start_row=row,start_column=12,end_row=row,end_column=13)
        _cell(ws,row,12,f"Saldo neto: {saldo}",ROJ if saldo>=3 else VER,BLA,True,9,"center")
        row+=1

        # Consecuencia
        for a,b,txt in ESCALA:
            if a<=saldo<=b:
                ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=L)
                _cell(ws,row,1,f"Art. 5 — {txt}",ROJP if saldo>=10 else ORAP,ROJ if saldo>=10 else ORA,True,9,wrap=True)
                row+=1; break
        row+=1
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=L)
        _cell(ws,row,1,"Director/a: _________________________________    Firma: _______________    Fecha: ___/___/______",GRP,NEG,False,8)

        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    # ── INSTRUMENTO 002 ──────────────────────────────────────────────────────
    def generar_consolidado_mensual(self,filas_por_mes,anio,cfg,nombre_docente,mes=None):
        wb=Workbook(); ws=wb.active; ws.title="Instrumento No. 002"

        # Columnas: A=Mes, B-D=Matrícula(M/H/Tot), E-G=Dem.Sexo, H-L=Dem.Causales,
        #           M-O=Red.Sexo, P-S=Red.Opción, T-V=Reconoc(M/H/Tot)
        # Total: 22 columnas (A..V)
        NCOLS = 22
        ws.column_dimensions["A"].width=13
        for i in range(2,NCOLS+1):
            ws.column_dimensions[get_column_letter(i)].width=7

        row=1

        # Títulos
        _hdr(ws,row,1,NCOLS,"MINISTERIO DE EDUCACIÓN, CIENCIA Y TECNOLOGÍA — MINEDUCYT",AZ,BLA,12); row+=1
        _hdr(ws,row,1,NCOLS,"Instrumento No. 002 — Registro Consolidado Mensual de Deméritos, Redenciones y Reconocimientos",AZL,BLA,10); ws.row_dimensions[row].height=22; row+=1
        _hdr(ws,row,1,NCOLS,"Memorándum N.° 06-2025 — Reglamento para la Promoción de la Cortesía Escolar",AZP,AZ,9,False); row+=1

        # Datos CE
        _hdr(ws,row,1,NCOLS,"DATOS DEL CENTRO EDUCATIVO",AZ,BLA,9); row+=1
        info=[("C.E.:",cfg.get("NOMBRE_INSTITUCION","—")),("Código:",cfg.get("CODIGO_CE","—")),
              ("Depto:",cfg.get("DEPARTAMENTO_CE","—")),("Municipio:",cfg.get("MUNICIPIO_CE","—")),
              ("Docente:",nombre_docente),
              ("Período:",("Mes "+["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][mes] if mes else "Anual")+f" {anio}")]
        col=1
        for i,(lbl,val) in enumerate(info):
            if i>0 and i%3==0: row+=1; col=1
            _cell(ws,row,col,lbl,AZP,AZ,True,8,"right"); col+=1
            ec=min(col+2,NCOLS)
            ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=ec)
            _cell(ws,row,col,val,BLA,NEG,False,8); col=ec+1
        row+=1

        # Encabezado tabla
        _hdr(ws,row,1,NCOLS,"TABLA DE DATOS — COLUMNAS 11 A 18",AZ,BLA,9); row+=1

        # Nivel 1 — grupos (columnas exactas sumando 22)
        # A=1, B-D=2-4, E-G=5-7, H-L=8-12, M-O=13-15, P-S=16-19, T-V=20-22
        g1=[(1,1,"11. Mes"),(2,4,"12. Matrícula"),(5,7,"13. Deméritos Sexo"),
            (8,12,"14. Deméritos Causales"),(13,15,"15. Redenciones Sexo"),
            (16,19,"16. Redenciones Opción"),(20,22,"17. Reconocimientos")]
        for sc,ec,txt in g1:
            _hdr(ws,row,sc,ec,txt,AZL,BLA,8)
        row+=1

        # Nivel 2 — sub-columnas (22 exactas)
        sub=[("Mes",AZP),("M",AZP),("H",AZP),("Tot",AZP),
             ("M",ROJP),("H",ROJP),("Tot",ROJP),
             ("A",ROJP),("B",ROJP),("C",ROJP),("D",ROJP),("Tot",ROJP),
             ("M",VERP),("H",VERP),("Tot",VERP),
             ("A",VERP),("B",VERP),("C",VERP),("Tot",VERP),
             ("M",ORAP),("H",ORAP),("Tot",ORAP)]
        for ci,(lbl,bg) in enumerate(sub,1):
            _cell(ws,row,ci,lbl,bg,NEG,True,8,"center")
        row+=1

        # Índice de datos por mes
        idx={r.get("mes_num",r.get("mes",0)):r for r in filas_por_mes}

        data_start=row
        for mes_n in range(1,12):
            r=idx.get(mes_n,{})
            bg4=GRP if mes_n%2==0 else BLA
            mm=int(r.get("matricula_m",0) or 0)
            mh=int(r.get("matricula_h",0) or 0)
            dm=int(r.get("demeritos_m",0) or 0)
            dh=int(r.get("demeritos_h",0) or 0)
            da=int(r.get("d_A",0) or 0); db=int(r.get("d_B",0) or 0)
            dc=int(r.get("d_C",0) or 0); dd=int(r.get("d_D",0) or 0)
            rm=int(r.get("redenciones_m",0) or 0)
            rh=int(r.get("redenciones_h",0) or 0)
            ra=int(r.get("r_A",0) or 0); rb=int(r.get("r_B",0) or 0); rc=int(r.get("r_C",0) or 0)
            recm=int(r.get("reconoc_m",0) or 0); rech=int(r.get("reconoc_h",0) or 0)

            vals=[MESES_ES.get(mes_n,str(mes_n)),
                  mm,mh,mm+mh,
                  dm,dh,dm+dh,
                  da,db,dc,dd,da+db+dc+dd,
                  rm,rh,rm+rh,
                  ra,rb,rc,ra+rb+rc,
                  recm,rech,recm+rech]

            _cell(ws,row,1,vals[0],AZP,AZ,True,9)
            for ci,v in enumerate(vals[1:],2):
                fg2=ROJ if 5<=ci<=12 else (VER if 13<=ci<=19 else (ORA if ci>=20 else NEG))
                _cell(ws,row,ci,v,bg4,fg2,False,9,"center")
            row+=1

        # Fila 18 — Totales con fórmulas SUM
        _cell(ws,row,1,"18. TOTAL ANUAL",AZ,BLA,True,9)
        for ci in range(2,NCOLS+1):
            cl=get_column_letter(ci)
            c=ws.cell(row=row,column=ci)
            if not isinstance(c,MergedCell):
                c.value=f"=SUM({cl}{data_start}:{cl}{row-1})"
                c.font=_font(True,BLA,10)
                c.fill=_fill(AZL)
                c.alignment=_align("center","center")
                c.border=_border(AZ,"medium")
        row+=2

        # Leyenda
        _hdr(ws,row,1,NCOLS,
             "CAUSALES (Art.3): A=No saludar  B=Omitir Por favor  C=Omitir Gracias  D=Tono irrespetuoso  |  "
             "REDENCIONES (Art.6): A=Semana cortesía  B=Limpieza  C=Campaña valores  |  "
             "RECONOCIMIENTOS (Art.7): A=Diploma  B=Mural",
             GRP,AZ,8,False); row+=1
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=NCOLS)
        _cell(ws,row,1,
              "Director/a: ________________________________    Firma: ______________    "
              "Docente: ________________________________    Fecha: ___/___/______",
              GRP,NEG,False,8)

        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf
