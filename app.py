"""app.py — Sistema Convivencia Escolar Web — MINEDUCYT Memorándum N.° 06-2025"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.chdir(os.path.dirname(os.path.abspath(__file__)))

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from functools import wraps
from datetime import datetime
from modules.database import DatabaseManager, CAUSALES_DEMERITO, OPCIONES_REDENCION, TIPOS_RECONOCIMIENTO, MESES
from modules.excel_generator import ExcelGenerator

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "convivencia_escolar_2026_mineducyt")
excel_gen = ExcelGenerator(output_dir="reports")

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "No autenticado"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("usuario", {}).get("rol") != "Administrador":
            return jsonify({"error": "Acceso denegado"}), 403
        return f(*args, **kwargs)
    return decorated

@app.route("/")
def index():
    return redirect(url_for("login") if "usuario" not in session else url_for("dashboard"))

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        d = request.get_json() or request.form
        u = DatabaseManager.autenticar_usuario((d.get("correo") or "").strip(), (d.get("password") or "").strip())
        if u:
            session["usuario"] = u
            return jsonify({"ok": True, "rol": u["rol"]})
        return jsonify({"ok": False, "error": "Credenciales incorrectas."}), 401
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("login"))

@app.route("/dashboard")
@login_required
def dashboard(): return render_template("base.html", pagina="dashboard", usuario=session["usuario"])

@app.route("/estudiantes")
@login_required
def estudiantes(): return render_template("base.html", pagina="estudiantes", usuario=session["usuario"])

@app.route("/registrar/<clase>")
@login_required
def registrar(clase):
    if clase not in ("DEMERITO","REDENCION","RECONOCIMIENTO"): return redirect(url_for("dashboard"))
    return render_template("base.html", pagina="registrar", clase=clase, usuario=session["usuario"])

@app.route("/consolidado")
@login_required
def consolidado(): return render_template("base.html", pagina="consolidado", usuario=session["usuario"])

@app.route("/usuarios")
@login_required
def usuarios(): return render_template("base.html", pagina="usuarios", usuario=session["usuario"])

@app.route("/api/dashboard")
@login_required
def api_dashboard():
    mes  = int(request.args.get("mes",  datetime.now().month))
    anio = int(request.args.get("anio", datetime.now().year))
    uid  = session["usuario"]["id_usuario"]
    rol  = session["usuario"].get("rol", "")
    try:
        # Docentes solo ven alertas de sus secciones asignadas
        if rol == "Administrador":
            grados_filtro = None
        else:
            grados_filtro = DatabaseManager.obtener_grados_usuario(uid)
        return jsonify({
            "stats":   DatabaseManager.obtener_estadisticas_dashboard(mes, anio, grados_filtro),
            "alertas": DatabaseManager.obtener_estudiantes_en_alerta(mes, anio, grados_filtro)
        })
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/estudiantes")
@login_required
def api_estudiantes():
    q        = request.args.get("q","").strip()
    id_grado = request.args.get("id_grado","").strip()
    uid      = session["usuario"]["id_usuario"]
    try: return jsonify(DatabaseManager.listar_estudiantes_por_usuario(uid, q, int(id_grado) if id_grado else None))
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/estudiantes", methods=["POST"])
@login_required
def api_crear_estudiante():
    try: return jsonify({"ok": True, "id": DatabaseManager.crear_estudiante(request.get_json())})
    except Exception as e: return jsonify({"error": str(e)}), 400

@app.route("/api/estudiantes/<int:id_est>")
@login_required
def api_estudiante_detalle(id_est):
    mes = int(request.args.get("mes", datetime.now().month))
    anio = int(request.args.get("anio", datetime.now().year))
    try:
        return jsonify({"estudiante": DatabaseManager.obtener_estudiante(id_est),
                        "tarjeta": DatabaseManager.listar_tarjeta_estudiante(id_est, mes, anio),
                        "totales": DatabaseManager.obtener_totales_tarjeta(id_est, mes, anio) or {}})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/grados")
@login_required
def api_grados(): return jsonify(DatabaseManager.listar_grados())

@app.route("/api/registrar", methods=["POST"])
@login_required
def api_registrar():
    datos = request.get_json()
    datos["id_docente_registra"] = session["usuario"]["id_usuario"]
    datos.setdefault("mes_periodo", datetime.now().month)
    datos.setdefault("anio_periodo", datetime.now().year)
    f = datos.get("fecha","")
    if "/" in str(f):
        p = str(f).split("/"); datos["fecha"] = f"{p[2]}-{p[1]}-{p[0]}"
    try:
        id_reg = DatabaseManager.registrar_evento(datos)
        totales = DatabaseManager.obtener_totales_tarjeta(datos["id_estudiante"], datos["mes_periodo"], datos["anio_periodo"])
        umbral = int(DatabaseManager.get_config("UMBRAL_BOLETA") or 3)
        return jsonify({"ok": True, "id": id_reg,
                        "alerta_umbral": bool(totales and totales.get("saldo_neto",0) >= umbral),
                        "totales": totales or {}})
    except Exception as e: return jsonify({"error": str(e)}), 400

@app.route("/api/demeritos_activos/<int:id_est>")
@login_required
def api_demeritos_activos(id_est):
    mes = int(request.args.get("mes", datetime.now().month))
    anio = int(request.args.get("anio", datetime.now().year))
    return jsonify(DatabaseManager.obtener_demeritos_activos_estudiante(id_est, mes, anio))

@app.route("/api/consolidado")
@login_required
def api_consolidado():
    anio = int(request.args.get("anio", datetime.now().year))
    try: return jsonify(DatabaseManager.obtener_detalle_consolidado_por_mes(anio))
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/plantilla_estudiantes")
@login_required
def api_plantilla_estudiantes():
    """Descarga la plantilla Excel para importación de estudiantes."""
    import io, os
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        ws.title = "Estudiantes"

        az, azl, gris = "003876", "1565C0", "F5F7FA"
        borde = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
                       top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))

        ws.merge_cells("A1:J1")
        ws["A1"] = "SISTEMA DE GESTIÓN DE CONVIVENCIA ESCOLAR — INTI"
        ws["A1"].font = Font(name="Arial",bold=True,size=13,color="FFFFFF")
        ws["A1"].fill = PatternFill("solid",start_color=az)
        ws["A1"].alignment = Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:J2")
        ws["A2"] = "Complete los datos desde la fila 4. Campos obligatorios: NIE, Nombre, Apellido, Sexo, Sección, Turno."
        ws["A2"].font = Font(name="Arial",size=9,italic=True,color="555555")
        ws["A2"].fill = PatternFill("solid",start_color="EBF3FF")
        ws["A2"].alignment = Alignment(horizontal="left",vertical="center",indent=1)
        ws.row_dimensions[2].height = 18

        headers = [
            ("A","NIE *",20),("B","Nombre *",20),("C","Apellido *",20),
            ("D","Sexo *",10),("E","Sección *",14),("F","Turno *",16),
            ("G","Nombre Responsable",24),("H","Teléfono",14),
            ("I","Correo Responsable",26),("J","Notas",20)
        ]
        for col, hdr, width in headers:
            c = ws[f"{col}3"]
            c.value = hdr
            c.font = Font(name="Arial",bold=True,size=10,color="FFFFFF")
            c.fill = PatternFill("solid",start_color=azl)
            c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            c.border = borde
            ws.column_dimensions[col].width = width
        ws.row_dimensions[3].height = 36

        for row in range(4,104):
            for col in "ABCDEFGHIJ":
                c = ws[f"{col}{row}"]
                c.font = Font(name="Arial",size=10)
                c.border = borde
                if row%2==0: c.fill = PatternFill("solid",start_color=gris)

        # Validaciones
        dv_s = DataValidation(type="list",formula1='"M,H"')
        ws.add_data_validation(dv_s); dv_s.sqref="D4:D103"
        dv_t = DataValidation(type="list",formula1='"Matutino,Vespertino,Nocturno"')
        ws.add_data_validation(dv_t); dv_t.sqref="F4:F103"
        dv_sec = DataValidation(type="list",formula1="Secciones!$A$2:$A$60")
        ws.add_data_validation(dv_sec); dv_sec.sqref="E4:E103"
        ws.freeze_panes = "A4"

        # Hoja Secciones
        ws2 = wb.create_sheet("Secciones")
        ws2["A1"] = "SECCIÓN"; ws2["B1"] = "ESPECIALIDAD"
        for c in [ws2["A1"],ws2["B1"]]:
            c.font=Font(name="Arial",bold=True,color="FFFFFF")
            c.fill=PatternFill("solid",start_color=az)
        ws2.column_dimensions["A"].width=14
        ws2.column_dimensions["B"].width=40

        grados = DatabaseManager.listar_grados()
        esp = {"DS":"Desarrollo de Software","MI":"Mecánica Industrial",
               "ECA":"Electrónica y Control Automatizado","ITSI":"Infraestructura TI",
               "SEER":"Sistemas Eléctricos y Energías Renovables","SE":"Sistemas Eléctricos",
               "MA":"Bachillerato General","FMA":"Matemáticas Avanzadas"}
        for i,g in enumerate(grados,2):
            codigo = (g["nombre"]+g["seccion"]).replace(" ","").replace("°","")
            c1 = ws2.cell(row=i,column=1,value=codigo)
            c1.font=Font(name="Arial",size=10); c1.border=borde
            if i%2==0: c1.fill=PatternFill("solid",start_color=gris)
            desc = next((v for k,v in esp.items() if codigo.upper().startswith(k)),"")
            c2 = ws2.cell(row=i,column=2,value=desc)
            c2.font=Font(name="Arial",size=10); c2.border=borde
            if i%2==0: c2.fill=PatternFill("solid",start_color=gris)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Plantilla_Importacion_Estudiantes.xlsx")
    except Exception as e:
        return jsonify({"error":str(e)}),500

@app.route("/api/importar_estudiantes", methods=["POST"])
@login_required
@admin_required
def api_importar_estudiantes():
    """Importa estudiantes desde archivo CSV o Excel."""
    import io, csv
    try:
        archivo = request.files.get("archivo")
        if not archivo:
            return jsonify({"error": "No se recibió archivo"}), 400

        nombre = archivo.filename.lower()
        filas = []

        if nombre.endswith(".csv"):
            contenido = archivo.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(contenido))
            for row in reader:
                filas.append(row)

        elif nombre.endswith(".xlsx") or nombre.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in ws[3]]
                for row in ws.iter_rows(min_row=4, values_only=True):
                    if not any(row):
                        continue
                    filas.append(dict(zip(headers, row)))
            except Exception as e:
                return jsonify({"error": f"Error leyendo Excel: {e}"}), 400
        else:
            return jsonify({"error": "Formato no soportado. Use .xlsx o .csv"}), 400

        # Obtener mapa de secciones nombre+seccion → id_grado
        grados = DatabaseManager.listar_grados()
        # Mapa por código de sección (ej: DS1A)
        grado_map = {}
        for g in grados:
            codigo = (g["nombre"] + g["seccion"]).replace(" ","").replace("°","")
            grado_map[codigo.upper()] = g["id_grado"]
            # También indexar por nombre+seccion directo
            grado_map[(g["nombre"]+g["seccion"]).upper()] = g["id_grado"]

        importados = 0
        errores = []

        for i, fila in enumerate(filas, start=4):
            # Leer columnas exactas de la tabla estudiantes
            nie      = str(fila.get("nie") or "").strip()
            nombre   = str(fila.get("nombre") or "").strip()
            apellido = str(fila.get("apellido") or "").strip()
            sexo     = str(fila.get("sexo") or "").strip().upper()
            id_grado_raw = str(fila.get("id_grado") or "").strip()
            turno    = str(fila.get("turno") or "Matutino").strip()
            id_centro= str(fila.get("id_centro") or "1").strip()
            resp     = str(fila.get("nombre_responsable") or "").strip()
            correo   = str(fila.get("correo_responsable") or "").strip()
            tel      = str(fila.get("telefono_responsable") or "").strip()
            activo   = str(fila.get("activo") or "1").strip()

            # Validar campos obligatorios
            if not nie or not nombre or not apellido:
                if any([nie, nombre, apellido]):
                    errores.append(f"Fila {i}: Faltan campos obligatorios (nie, nombre o apellido)")
                continue

            if sexo not in ("M", "H"):
                errores.append(f"Fila {i} ({nie}): sexo inválido '{sexo}' — use M o H")
                continue

            # id_grado puede ser número directo o código de sección
            id_grado = None
            if id_grado_raw.isdigit():
                id_grado = int(id_grado_raw)
            else:
                id_grado = grado_map.get(id_grado_raw.upper())
            if not id_grado:
                errores.append(f"Fila {i} ({nie}): id_grado '{id_grado_raw}' no encontrado")
                continue

            if turno not in ("Matutino", "Vespertino", "Nocturno"):
                turno = "Matutino"

            try:
                id_centro_int = int(id_centro) if id_centro.isdigit() else 1
            except:
                id_centro_int = 1

            try:
                activo_int = int(activo) if activo in ("0","1") else 1
            except:
                activo_int = 1

            try:
                conn = DatabaseManager.get_connection()
                try:
                    DatabaseManager.execute_query(conn, """
                        INSERT INTO estudiantes
                            (nie, nombre, apellido, sexo, id_grado, turno, id_centro,
                             nombre_responsable, correo_responsable, telefono_responsable, activo)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON DUPLICATE KEY UPDATE
                            nombre=VALUES(nombre),
                            apellido=VALUES(apellido),
                            sexo=VALUES(sexo),
                            id_grado=VALUES(id_grado),
                            turno=VALUES(turno),
                            id_centro=VALUES(id_centro),
                            nombre_responsable=VALUES(nombre_responsable),
                            correo_responsable=VALUES(correo_responsable),
                            telefono_responsable=VALUES(telefono_responsable),
                            activo=VALUES(activo)
                    """, (nie, nombre, apellido, sexo, id_grado, turno, id_centro_int,
                          resp or None, correo or None, tel or None, activo_int))
                    conn.commit()
                    importados += 1
                finally:
                    conn.close()
            except Exception as e:
                errores.append(f"Fila {i} ({nie}): {str(e)[:80]}")

        return jsonify({
            "ok": True,
            "importados": importados,
            "errores": errores,
            "total_filas": len(filas)
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/consolidado/filtrado")
@login_required
@admin_required
def api_consolidado_filtrado():
    anio      = int(request.args.get("anio", datetime.now().year))
    id_grado  = request.args.get("id_grado") or None
    id_docente= request.args.get("id_docente") or None
    try:
        return jsonify(DatabaseManager.obtener_consolidado_filtrado(anio, id_grado, id_docente))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/estudiantes/<int:id_est>", methods=["PUT"])
@login_required
def api_actualizar_estudiante(id_est):
    datos = request.get_json()
    try:
        DatabaseManager.actualizar_estudiante(id_est, datos)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/estudiantes/<int:id_est>", methods=["DELETE"])
@login_required
@admin_required
def api_eliminar_estudiante(id_est):
    try:
        DatabaseManager.eliminar_estudiante(id_est)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/estudiantes/<int:id_est>/datos")
@login_required
def api_datos_estudiante(id_est):
    try:
        est = DatabaseManager.obtener_estudiante(id_est)
        if not est: return jsonify({"error": "No encontrado"}), 404
        return jsonify(est)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/usuarios/<int:uid>", methods=["PUT"])
@login_required
@admin_required
def api_actualizar_usuario(uid):
    d = request.get_json()
    try:
        conn = DatabaseManager.get_connection()
        try:
            DatabaseManager.execute_query(conn,
                "UPDATE usuarios SET nombre_completo=%s, correo=%s, id_rol=%s, activo=%s WHERE id_usuario=%s",
                (d["nombre"], d["correo"], int(d["id_rol"]), int(d["activo"]), uid))
            conn.commit()
        finally:
            conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/usuarios/<int:uid>", methods=["DELETE"])
@login_required
@admin_required
def api_eliminar_usuario(uid):
    if uid == session["usuario"]["id_usuario"]:
        return jsonify({"error": "No puede eliminar su propia cuenta"}), 400
    try:
        DatabaseManager.eliminar_usuario(uid)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/docentes")
@login_required
@admin_required
def api_docentes():
    """Lista usuarios con rol Docente para filtros."""
    try:
        todos = DatabaseManager.listar_usuarios()
        return jsonify([u for u in todos if u.get("rol_nombre") != "Administrador" or True])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/plantilla_usuarios")
@login_required
@admin_required
def api_plantilla_usuarios():
    """Genera y descarga la plantilla Excel para importación de usuarios."""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        az, azl, gris = "003876", "1565C0", "F5F7FA"
        borde = Border(
            left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC')
        )

        # Fila 1: Título
        ws.merge_cells("A1:G1")
        ws["A1"] = "PLANTILLA DE IMPORTACIÓN DE USUARIOS — SISTEMA INTI"
        ws["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", start_color=az)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Fila 2: Instrucciones
        ws.merge_cells("A2:G2")
        ws["A2"] = "Deje id_usuario vacío. contrasena = contraseña inicial (mín. 6 caracteres). activo: 1=Activo, 0=Inactivo."
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color="555555")
        ws["A2"].fill = PatternFill("solid", start_color="EBF3FF")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18

        # Fila 3: Encabezados exactos de la tabla usuarios
        cols = [
            ("A", "id_usuario",      "Dejar vacío (auto)",                        14),
            ("B", "nombre_completo", "Nombre completo del docente/usuario",        28),
            ("C", "correo",          "Correo institucional (ej: docente@inti.edu.sv)", 30),
            ("D", "contrasena",      "Contraseña inicial (mín. 6 caracteres)",     20),
            ("E", "id_rol",          "1=Administrador  2=Docente",                 14),
            ("F", "activo",          "1=Activo  0=Inactivo",                       10),
            ("G", "primer_login",    "1=Debe cambiar pwd  0=No (default: 1)",       18),
        ]
        for col, hdr, nota, width in cols:
            c = ws[f"{col}3"]
            c.value = hdr
            c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", start_color=azl)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = borde
            ws.column_dimensions[col].width = width
        ws.row_dimensions[3].height = 36

        # 100 filas de datos con valores por defecto
        for row in range(4, 104):
            for col in "ABCDEFG":
                c = ws[f"{col}{row}"]
                c.font = Font(name="Arial", size=10)
                c.border = borde
                if row % 2 == 0:
                    c.fill = PatternFill("solid", start_color=gris)
            ws[f"E{row}"] = 2  # id_rol default = Docente
            ws[f"F{row}"] = 1  # activo default = 1
            ws[f"G{row}"] = 1  # primer_login default = 1

        # Validaciones
        dv_rol = DataValidation(type="list", formula1='"1,2"')
        ws.add_data_validation(dv_rol); dv_rol.sqref = "E4:E103"
        dv_act = DataValidation(type="list", formula1='"1,0"')
        ws.add_data_validation(dv_act); dv_act.sqref = "F4:F103"
        dv_pl  = DataValidation(type="list", formula1='"1,0"')
        ws.add_data_validation(dv_pl);  dv_pl.sqref  = "G4:G103"
        ws.freeze_panes = "A4"

        # Hoja de referencia roles
        ws2 = wb.create_sheet("Referencia")
        ws2.merge_cells("A1:C1")
        ws2["A1"] = "REFERENCIA — Roles del sistema"
        ws2["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws2["A1"].fill = PatternFill("solid", start_color=az)
        ws2["A1"].alignment = Alignment(horizontal="center")
        for col, h in [("A","id_rol"),("B","Nombre del rol"),("C","Descripción")]:
            c = ws2[f"{col}2"]
            c.value = h; c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", start_color=azl); c.border = borde
        for i, (rid, rnom, rdesc) in enumerate([(1,"Administrador","Acceso completo al sistema"),
                                                  (2,"Docente","Registro y consulta de sus secciones")], 3):
            for j, v in enumerate([rid, rnom, rdesc], 1):
                c = ws2.cell(row=i, column=j, value=v)
                c.font = Font(name="Arial", size=10); c.border = borde
                if i % 2 == 0: c.fill = PatternFill("solid", start_color=gris)
        ws2.column_dimensions["A"].width=10
        ws2.column_dimensions["B"].width=20
        ws2.column_dimensions["C"].width=40

        # Hoja Ejemplo
        ws3 = wb.create_sheet("Ejemplo")
        ws3.merge_cells("A1:G1")
        ws3["A1"] = "EJEMPLO — No modificar esta hoja"
        ws3["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws3["A1"].fill = PatternFill("solid", start_color="C62828")
        ws3["A1"].alignment = Alignment(horizontal="center")
        ej_hdrs = ["id_usuario","nombre_completo","correo","contrasena","id_rol","activo","primer_login"]
        for ci, h in enumerate(ej_hdrs, 1):
            c = ws3.cell(row=2, column=ci, value=h)
            c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", start_color=azl); c.border = borde
        ejemplos = [
            ["","María García","mgarcia@inti.edu.sv","Docente2024!",2,1,1],
            ["","Carlos López","clopez@inti.edu.sv","Docente2024!",2,1,1],
            ["","Ana Martínez","amartinez@inti.edu.sv","Admin2024!",1,1,0],
        ]
        for ri, row_data in enumerate(ejemplos, 3):
            for ci, v in enumerate(row_data, 1):
                c = ws3.cell(row=ri, column=ci, value=v)
                c.font = Font(name="Arial", size=10); c.border = borde
                if ri%2==0: c.fill = PatternFill("solid", start_color=gris)
        for col in "ABCDEFG":
            ws3.column_dimensions[col].width = ws.column_dimensions[col].width

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name="Plantilla_Usuarios.xlsx")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/importar_usuarios", methods=["POST"])
@login_required
@admin_required
def api_importar_usuarios():
    """Importa usuarios desde Excel o CSV."""
    import io, csv
    try:
        archivo = request.files.get("archivo")
        if not archivo:
            return jsonify({"error": "No se recibió archivo"}), 400
        nombre = archivo.filename.lower()
        filas = []

        if nombre.endswith(".csv"):
            contenido = archivo.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(contenido))
            for row in reader:
                filas.append(row)
        elif nombre.endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in ws[3]]
                for row in ws.iter_rows(min_row=4, values_only=True):
                    if not any(row): continue
                    filas.append(dict(zip(headers, row)))
            except Exception as e:
                return jsonify({"error": f"Error leyendo Excel: {e}"}), 400
        else:
            return jsonify({"error": "Use .xlsx o .csv"}), 400

        importados, errores = 0, []

        for i, fila in enumerate(filas, start=4):
            nombre_c  = str(fila.get("nombre_completo") or "").strip()
            correo    = str(fila.get("correo") or "").strip()
            pwd       = str(fila.get("contrasena") or "").strip()
            id_rol    = str(fila.get("id_rol") or "2").strip()
            activo    = str(fila.get("activo") or "1").strip()
            primer_l  = str(fila.get("primer_login") or "1").strip()

            if not nombre_c or not correo or not pwd:
                if nombre_c or correo:
                    errores.append(f"Fila {i}: Faltan nombre_completo, correo o contrasena")
                continue
            if len(pwd) < 6:
                errores.append(f"Fila {i} ({correo}): Contraseña muy corta (mín. 6 caracteres)")
                continue
            try:
                id_rol_int   = int(id_rol) if id_rol in ("1","2") else 2
                activo_int   = int(activo) if activo in ("0","1") else 1
                primer_l_int = int(primer_l) if primer_l in ("0","1") else 1
                DatabaseManager.crear_usuario(nombre_c, correo, pwd, id_rol_int)
                # Actualizar activo y primer_login
                conn = DatabaseManager.get_connection()
                try:
                    DatabaseManager.execute_query(conn,
                        "UPDATE usuarios SET activo=%s, primer_login=%s WHERE correo=%s",
                        (activo_int, primer_l_int, correo))
                    conn.commit()
                finally:
                    conn.close()
                importados += 1
            except Exception as e:
                errores.append(f"Fila {i} ({correo}): {str(e)[:80]}")

        return jsonify({"ok": True, "importados": importados,
                        "errores": errores, "total_filas": len(filas)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/usuarios")
@login_required
@admin_required
def api_usuarios(): return jsonify(DatabaseManager.listar_usuarios())

@app.route("/api/usuarios", methods=["POST"])
@login_required
@admin_required
def api_crear_usuario():
    d = request.get_json()
    try: DatabaseManager.crear_usuario(d["nombre"],d["correo"],d["password"],int(d["id_rol"])); return jsonify({"ok":True})
    except Exception as e: return jsonify({"error":str(e)}), 400

@app.route("/api/usuarios/<int:uid>/reset", methods=["POST"])
@login_required
@admin_required
def api_reset_password(uid):
    DatabaseManager.resetear_contrasena(uid, request.get_json()["nueva"]); return jsonify({"ok":True})

@app.route("/api/usuarios/<int:uid>/toggle", methods=["POST"])
@login_required
@admin_required
def api_toggle_usuario(uid):
    DatabaseManager.togglear_usuario(uid, request.get_json()["activo"]); return jsonify({"ok":True})

@app.route("/api/usuarios/<int:uid>/grados", methods=["GET"])
@login_required
def api_get_grados_usuario(uid):
    try: return jsonify(DatabaseManager.listar_grados_con_asignacion(uid))
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/usuarios/<int:uid>/grados", methods=["POST"])
@login_required
@admin_required
def api_set_grados_usuario(uid):
    try:
        ids = request.get_json().get("grados", [])
        DatabaseManager.asignar_grados_usuario(uid, ids)
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 400

@app.route("/api/mis_grados")
@login_required
def api_mis_grados():
    uid  = session["usuario"]["id_usuario"]
    rol  = session["usuario"].get("rol", "")
    try:
        todos = DatabaseManager.listar_grados()
        if not todos:
            return jsonify([])

        # Admin: devuelve todos con asignado=1 (ve todo)
        if rol == "Administrador":
            return jsonify([{
                "id_grado": g["id_grado"],
                "nombre":   g["nombre"],
                "seccion":  g["seccion"],
                "nivel":    g.get("nivel", ""),
                "asignado": 1
            } for g in todos])

        # Docente: obtener solo los grados asignados en usuario_grados
        try:
            from mysql.connector import Error as MErr
            conn = DatabaseManager.get_connection()
            try:
                rows = DatabaseManager.execute_query(conn,
                    "SELECT id_grado FROM usuario_grados WHERE id_usuario=%s",
                    (uid,), fetch=True)
                asignados = set(r["id_grado"] for r in rows)
            finally:
                conn.close()
        except Exception:
            asignados = set(g["id_grado"] for g in todos)

        # Si el docente no tiene ninguna asignación, ver todos
        if not asignados:
            asignados = set(g["id_grado"] for g in todos)

        return jsonify([{
            "id_grado": g["id_grado"],
            "nombre":   g["nombre"],
            "seccion":  g["seccion"],
            "nivel":    g.get("nivel", ""),
            "asignado": 1 if g["id_grado"] in asignados else 0
        } for g in todos])

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/test_grados")
@login_required
def api_test_grados():
    """Endpoint de diagnóstico — verificar en navegador: /api/test_grados"""
    uid = session["usuario"]["id_usuario"]
    resultado = {"uid": uid, "pasos": []}
    try:
        grados = DatabaseManager.listar_grados()
        resultado["pasos"].append(f"listar_grados OK: {len(grados)} grados")
    except Exception as e:
        resultado["pasos"].append(f"listar_grados ERROR: {e}")
        return jsonify(resultado)
    try:
        asig = DatabaseManager.obtener_grados_usuario(uid)
        resultado["pasos"].append(f"obtener_grados_usuario OK: {asig}")
    except Exception as e:
        resultado["pasos"].append(f"obtener_grados_usuario ERROR: {e}")
    try:
        from mysql.connector import Error as MErr
        conn = DatabaseManager.get_connection()
        cur = conn.cursor()
        cur.execute("SHOW TABLES LIKE 'usuario_grados'")
        existe = cur.fetchone()
        resultado["pasos"].append(f"tabla usuario_grados: {'EXISTE' if existe else 'NO EXISTE'}")
        conn.close()
    except Exception as e:
        resultado["pasos"].append(f"check tabla ERROR: {e}")
    return jsonify(resultado)

@app.route("/api/cambiar_contrasena", methods=["POST"])
@login_required
def api_cambiar_contrasena():
    """Permite al docente cambiar su contraseña en el primer acceso o cuando lo desee."""
    d = request.get_json()
    uid = session["usuario"]["id_usuario"]
    nueva = (d.get("nueva") or "").strip()
    if len(nueva) < 6:
        return jsonify({"error": "La contraseña debe tener al menos 6 caracteres"}), 400
    try:
        from modules.database import DatabaseManager as DB
        nuevo_hash = DB.hash_password(nueva)
        conn = DB.get_connection()
        try:
            DB.execute_query(conn,
                "UPDATE usuarios SET contrasena_hash=%s, primer_login=0 WHERE id_usuario=%s",
                (nuevo_hash, uid))
            conn.commit()
        finally:
            conn.close()
        # Actualizar sesión
        session["usuario"]["primer_login"] = 0
        session.modified = True
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/roles")
@login_required
def api_roles(): return jsonify(DatabaseManager.listar_roles())

@app.route("/api/excel/tarjeta/<int:id_est>")
@login_required
def api_excel_tarjeta(id_est):
    mes = int(request.args.get("mes", datetime.now().month))
    anio = int(request.args.get("anio", datetime.now().year))
    try:
        est = DatabaseManager.obtener_estudiante(id_est)
        buf = excel_gen.generar_tarjeta_demerito(est,
              DatabaseManager.listar_tarjeta_estudiante(id_est, mes, anio),
              DatabaseManager.obtener_totales_tarjeta(id_est, mes, anio) or {},
              mes, anio, DatabaseManager.get_config_ce())
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"Instrumento001_{est.get('nie',id_est)}_{anio}_{mes:02d}.xlsx")
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/excel/consolidado")
@login_required
def api_excel_consolidado():
    anio       = int(request.args.get("anio", datetime.now().year))
    mes        = request.args.get("mes") or None
    id_grado   = request.args.get("id_grado") or None
    id_docente = request.args.get("id_docente") or None
    try:
        # Obtener datos — filtrados si se especifica grado/docente
        if id_grado or id_docente:
            filas = DatabaseManager.obtener_consolidado_filtrado(anio, id_grado, id_docente)
        else:
            filas = DatabaseManager.obtener_detalle_consolidado_por_mes(anio)

        # Filtrar por mes específico si se solicitó
        if mes:
            mes_int = int(mes)
            filas = [r for r in filas if int(r.get("mes_num", r.get("mes", 0))) == mes_int]

        cfg = DatabaseManager.get_config_ce()
        nombre_doc = session["usuario"]["nombre_completo"]

        # Nombre del archivo descriptivo
        sufijo = ""
        if mes:
            meses_es = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                        "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
            sufijo += f"_{meses_es[int(mes)]}"
        if id_grado:
            sufijo += f"_Grado{id_grado}"
        if id_docente:
            sufijo += f"_Docente{id_docente}"

        buf = excel_gen.generar_consolidado_mensual(filas, anio, cfg, nombre_doc,
                                                     mes=int(mes) if mes else None)
        fname = f"Instrumento002_Consolidado_{anio}{sufijo}.xlsx"
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=fname)
    except Exception as e: return jsonify({"error": str(e)}), 500

def _ensure_usuario_grados():
    """Crea la tabla usuario_grados y columna primer_login si no existen (migración automática)."""
    try:
        conn = DatabaseManager.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS usuario_grados (
                    id         INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                    id_usuario INT UNSIGNED NOT NULL,
                    id_grado   INT UNSIGNED NOT NULL,
                    UNIQUE KEY uq_usr_grado (id_usuario, id_grado),
                    CONSTRAINT fk_ug_usr   FOREIGN KEY (id_usuario)
                        REFERENCES usuarios(id_usuario) ON DELETE CASCADE,
                    CONSTRAINT fk_ug_grado FOREIGN KEY (id_grado)
                        REFERENCES grados(id_grado) ON DELETE CASCADE
                )
            """)
            conn.commit()
            # Agregar columna primer_login si no existe
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN primer_login TINYINT(1) DEFAULT 1")
                conn.commit()
            except Exception:
                pass  # Ya existe
            cursor.close()
        finally:
            conn.close()
    except Exception as e:
        print(f"  Aviso usuario_grados: {e}")

if __name__ == "__main__":
    print("\n  Iniciando servidor web en:")
    print("    http://localhost:5000")
    print("    http://" + __import__("socket").gethostname() + ":5000  (red local)")
    print("\n  Abra su navegador y vaya a: http://localhost:5000")
    print("  Para detener el servidor presione Ctrl+C\n")
    try:
        DatabaseManager.init_pool()
        _ensure_usuario_grados()
        print("  Base de datos conectada correctamente.\n")
    except Exception as e:
        print(f"\n  ERROR conectando a MySQL: {e}")
        print("  Verifique que XAMPP MySQL esté activo (botón verde).\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
if __name__ == "__main__":
    import os
    try:
        DatabaseManager.init_pool()
        _ensure_usuario_grados()
        print("Base de datos conectada correctamente.")
    except Exception as e:
        print(f"Error DB: {e}")

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)